from __future__ import annotations

import csv
import json
import logging
import os
import tempfile
from collections.abc import Callable, Iterable
from itertools import chain
from pathlib import Path
from threading import Event
from typing import Any

from openpyxl import Workbook

from arenyxa.domain.errors import ArenyxaError
from arenyxa.infrastructure.atomic_io import fsync_existing_file
from arenyxa.infrastructure.database import SQLiteStore

LOGGER = logging.getLogger(__name__)


class ExportService:
    def __init__(self, store: SQLiteStore) -> None:
        self.store = store

    def export_run(
        self,
        run_id: str,
        destination: Path,
        format_name: str,
        *,
        cancel: Event | None = None,
        progress: Callable[[int], None] | None = None,
    ) -> int:
        destination.parent.mkdir(parents=True, exist_ok=True)
        normalized = format_name.lower()
        writers = {
            "csv": self._csv,
            "jsonl": self._jsonl,
            "ndjson": self._jsonl,
            "json": self._json,
            "xlsx": self._xlsx,
            "excel": self._xlsx,
        }
        writer = writers.get(normalized)
        if writer is None:
            raise ValueError(f"unsupported export format: {format_name}")

                                                                                        
                                                                                          
                                                                                     
        fd, raw_temp = tempfile.mkstemp(
            prefix=f".{destination.name}.", suffix=".tmp", dir=destination.parent
        )
        os.close(fd)
        temp_path = Path(raw_temp)
        try:
            rows = self.store.iter_results(run_id)
            count = writer(rows, temp_path, cancel, progress)
            if cancel and cancel.is_set():
                raise ArenyxaError("EXPORT_CANCELLED", "导出已取消。", domain="EXPORT")
                                                                                             
                                                                                             
                                                       
            fsync_existing_file(temp_path)
            os.replace(temp_path, destination)
            return count
        finally:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass

    @staticmethod
    def _rows(rows: Iterable[dict[str, Any]], cancel: Event | None) -> Iterable[dict[str, Any]]:
        for row in rows:
            if cancel and cancel.is_set():
                break
            yield row

    def _csv(
        self,
        rows: Iterable[dict[str, Any]],
        path: Path,
        cancel: Event | None,
        progress: Callable[[int], None] | None,
    ) -> int:
        iterator = iter(self._rows(rows, cancel))
        first = next(iterator, None)
        if first is None:
            path.write_text("", encoding="utf-8-sig")
            return 0
        fields = list(first)
        count = 0
        with path.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for row in chain((first,), iterator):
                writer.writerow({key: self._scalar(value) for key, value in row.items()})
                count += 1
                if progress and count % 250 == 0:
                    self._notify_progress(progress, count)
        return count

    def _jsonl(
        self,
        rows: Iterable[dict[str, Any]],
        path: Path,
        cancel: Event | None,
        progress: Callable[[int], None] | None,
    ) -> int:
        count = 0
        with path.open("w", encoding="utf-8", newline="\n") as stream:
            for row in self._rows(rows, cancel):
                stream.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
                count += 1
                if progress and count % 250 == 0:
                    self._notify_progress(progress, count)
        return count

    def _json(
        self,
        rows: Iterable[dict[str, Any]],
        path: Path,
        cancel: Event | None,
        progress: Callable[[int], None] | None,
    ) -> int:
        count = 0
        with path.open("w", encoding="utf-8", newline="\n") as stream:
            stream.write("[\n")
            first = True
            for row in self._rows(rows, cancel):
                if not first:
                    stream.write(",\n")
                stream.write(json.dumps(row, ensure_ascii=False, default=str))
                first = False
                count += 1
                if progress and count % 250 == 0:
                    self._notify_progress(progress, count)
            stream.write("\n]\n")
        return count

    def _xlsx(
        self,
        rows: Iterable[dict[str, Any]],
        path: Path,
        cancel: Event | None,
        progress: Callable[[int], None] | None,
    ) -> int:
        workbook = Workbook(write_only=True)
        sheet_index = 1
        sheet = workbook.create_sheet("Arenyxa Data")
        fields: list[str] = []
        sheet_rows = 0
        count = 0
        for row in self._rows(rows, cancel):
            if not fields:
                fields = list(row)
                if len(fields) > 16_384:
                    raise ArenyxaError(
                        "EXPORT_XLSX_TOO_MANY_COLUMNS",
                        "XLSX 单表最多支持 16384 列；请改用 JSONL/CSV 或减少字段。",
                        domain="EXPORT",
                    )
                sheet.append(fields)
                sheet_rows = 1
                                                                                         
                                                                                            
                                                          
            if sheet_rows >= 1_048_576:
                sheet_index += 1
                sheet = workbook.create_sheet(f"Arenyxa Data {sheet_index}")
                sheet.append(fields)
                sheet_rows = 1
            values = [self._scalar(row.get(field)) for field in fields]
            if any(isinstance(value, str) and len(value) > 32_767 for value in values):
                raise ArenyxaError(
                    "EXPORT_XLSX_CELL_TOO_LARGE",
                    "某个单元格超过 Excel 32767 字符限制；为避免静默截断，请改用 JSONL/JSON。",
                    domain="EXPORT",
                )
            sheet.append(values)
            sheet_rows += 1
            count += 1
            if progress and count % 250 == 0:
                self._notify_progress(progress, count)
        workbook.save(path)
        return count

    @staticmethod
    def _notify_progress(callback: Callable[[int], None], count: int) -> None:
        try:
            callback(count)
        except Exception:
                                                                                           
                                                                      
            LOGGER.exception("Export progress callback failed")

    @staticmethod
    def _scalar(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return value
